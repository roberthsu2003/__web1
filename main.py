from flask import Flask,render_template,redirect,url_for,session
import secrets

from flask_wtf import FlaskForm
from wtforms import StringField,PasswordField
from wtforms.validators import DataRequired

from api.index  import api



class User_Login(FlaskForm):
    name = StringField("使用者名稱",validators=[DataRequired()])
    password  = PasswordField("密碼",validators=[DataRequired()])


app = Flask(__name__)
app.config['SECRET_KEY'] = secrets.token_hex()
app.register_blueprint(api,url_prefix="/api")

@app.route("/")
def index():
    return render_template('index.html')

from openpyxl import load_workbook
@app.route('/login',methods=['GET','POST'])
def login():
    form = User_Login()
    context =  {
        'form':form
    }
    if form.is_submitted() and form.validate():
        username = form.name.data
        password  =  form.password.data
        wb = load_workbook(filename='static/others/login.xlsx')
        sheet_ranges = wb['登入']        
        for row in sheet_ranges.iter_rows():
            if row[0].value == username  and row[1].value ==  password:                
                session['login_name'] = username
                return redirect(url_for('index'))
        
        context['error'] = "帳號密碼有誤"        
        return render_template('login.html',**context)
        return f"使用者名稱:{form.name.data},使用者密碼:{form.password.data}"
    else:
        return render_template('login.html',**context)

@app.route('/logout')
def logout():
    session.pop('login_name',None)
    return redirect(url_for('index'))



